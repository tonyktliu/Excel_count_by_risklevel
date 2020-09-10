import csv
import openpyxl as xl;
import sys
from openpyxl.styles import Font

# ===============Variables for the script=================#
targetExcel = "Test_Output.xlsx"
prefixKeyword = ["None","Low","Medium","High","Critical"]
referenceName = "Risk"
referenceRow = 1
targetxlsx = 'temp.xlsx'

# ========================================================#

def conversion(source, target):
    wb = xl.Workbook()
    ws = wb.active

    with open(source, 'r', encoding='UTF-8') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save(target)


def riskCount(inputfile, outputfile):
    # opening the source excel file
    filename = inputfile
    wb1 = xl.load_workbook(filename)

    #Init the counters
    #counter = [0,0,0,0,0]
    KeywordLength = len(prefixKeyword)
    counter = []
    for x in range(0, KeywordLength):
        counter.append(0)

    # opening the destination excel file
    filename1 = outputfile
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2["Output"]

    sheets = wb1.sheetnames
    print("The sheets for processing include:", sheets)
    x = len(sheets)

    #Loop for all sheets.
    for z in range(x):
        ws1 = wb1.worksheets[z]

        # calculate total number of rows and
        # columns in the source excel file
        mr = ws1.max_row
        mc = ws1.max_column

        # Initialize a variable for checking the column position of reference value.
        referVal = 999
        for a in range(1, mc + 1):
            if ws1.cell(row=referenceRow, column=a).value == referenceName:
                referVal = a
                break

        if referVal == 999:
            print("ERROR: No Reference Name cell was found.")
            sys.exit(1)
        #print("referVal:", referVal)

        # Set header
        ws2['A1'] = "Risk Level"
        ws2['B1'] = "Number of Cases"

        # Set font style
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)

        # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            tempText = ws1.cell(row=i, column=referVal).value

            for x in range(0, KeywordLength):
            # Only compare with Keyword when the cell is NOT empty.
                if tempText:
                    if tempText.startswith(prefixKeyword[x]):
                        counter[x] += 1


    lastrowCounterSheet = len(ws2['A'])  # Check the last row of Column A for appending.

    for x in range(0, KeywordLength):
        ws2.cell(row=lastrowCounterSheet + x + 1, column=1).value = prefixKeyword[x]
        ws2.cell(row=lastrowCounterSheet + x + 1, column=2).value = counter[x]

    countertotal = 0
    lastrowCounterSheet = len(ws2['A'])  # Check the last row of Column A for appending.
    ws2.cell(row=lastrowCounterSheet + 1, column=1).value = "Total"
    for x in range(0, KeywordLength):
        countertotal = countertotal + counter[x]

    ws2.cell(row=lastrowCounterSheet + 1, column=2).value = countertotal

    # saving the destination excel file
    wb2.save(str(filename1))


def clearsheet(outputfile):
    wb2 = xl.load_workbook(outputfile)
    ws2 = wb2["Output"]
    for row in ws2['A1:Z999']:
        for cell in row:
            cell.value = None
    wb2.save(str(outputfile))


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("ERROR: Please enter a valid source filename.")
        sys.exit(1)

    try:
        print("Source filename: %s" % (sys.argv[1]))
        sourcecsv = sys.argv[1]

        conversion(sourcecsv, targetxlsx)
        clearsheet(targetExcel)
        riskCount(targetxlsx, targetExcel)

        print("The program has been completed. Please check the output file:", targetExcel)
    except:
        print("Unexpected error:", sys.exc_info()[0])
        raise
